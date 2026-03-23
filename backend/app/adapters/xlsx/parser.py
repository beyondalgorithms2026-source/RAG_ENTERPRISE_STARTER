from io import BytesIO
from typing import List

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from app.adapters.models import ParsedSourceDocument, ParsedSourcePart


def _cell_value_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_xlsx_bytes(content: bytes, file_name: str) -> ParsedSourceDocument:
    warnings = []
    parts = []
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    for part_index, sheet_name in enumerate(workbook.sheetnames):
        worksheet = workbook[sheet_name]
        lines: List[str] = []
        non_empty_cells = 0
        min_row = None
        max_row = None
        min_col = None
        max_col = None
        for row in worksheet.iter_rows():
            cell_values = []
            row_number = row[0].row if row else None
            for cell in row:
                value = _cell_value_text(cell.value)
                if not value:
                    continue
                non_empty_cells += 1
                min_row = cell.row if min_row is None else min(min_row, cell.row)
                max_row = cell.row if max_row is None else max(max_row, cell.row)
                min_col = cell.column if min_col is None else min(min_col, cell.column)
                max_col = cell.column if max_col is None else max(max_col, cell.column)
                cell_values.append(f"{cell.coordinate}={value}")
            if cell_values and row_number is not None:
                lines.append(f"Row {row_number}: " + " | ".join(cell_values))
        if not lines:
            continue
        range_ref = None
        if min_row is not None and max_row is not None and min_col is not None and max_col is not None:
            range_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        parts.append(
            ParsedSourcePart(
                part_type="sheet",
                part_index=part_index,
                title=sheet_name,
                locator_json={"sheet": sheet_name, "range": range_ref},
                content_text="\n".join(lines),
                provenance_json={"parser": "openpyxl", "file_name": file_name, "non_empty_cells": non_empty_cells},
            )
        )

    if not parts:
        warnings.append("No XLSX sheet text found.")

    return ParsedSourceDocument(
        source_type="xlsx",
        title=file_name,
        metadata={"file_name": file_name, "sheet_count": len(workbook.sheetnames), "sheet_names": list(workbook.sheetnames)},
        parts=parts,
        warnings=warnings,
    )
